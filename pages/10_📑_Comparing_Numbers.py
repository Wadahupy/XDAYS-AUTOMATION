import streamlit as st
import pandas as pd
from datetime import datetime
import os

st.set_page_config(page_title="Comparing Numbers", layout="wide")
st.title("📑 Worklist Comparison Checker")

st.markdown("""
Compare yesterday's and today's worklists to identify changes in customer contact information.
**Unique Identifiers**: CUST_ID + MOBILE_NO
""")

# Columns to compare
COLUMNS_TO_COMPARE = ["CUST_ID", "CUST_NAME", "OFFICE_PH", "HOME_PH", "MOBILE_NO"]

def load_and_validate_file(uploaded_file):
    """Load Excel file and validate required columns"""
    try:
        df = pd.read_excel(uploaded_file)
        # Normalize column names (uppercase, strip whitespace)
        df.columns = df.columns.str.strip().str.upper()
        
        # Check if required columns exist
        missing_cols = [col for col in COLUMNS_TO_COMPARE if col not in df.columns]
        if missing_cols:
            st.error(f"❌ Missing required columns: {', '.join(missing_cols)}")
            return None
        
        return df[COLUMNS_TO_COMPARE].copy()
    except Exception as e:
        st.error(f"❌ Error loading file: {str(e)}")
        return None


def compare_worklists(yesterday_df, today_df):
    """
    Compare two worklists and identify differences.
    
    Returns:
        - changes_df: DataFrame with all changes found
        - summary_stats: Dictionary with statistics
        - columns_changed_overall: Set of all columns that changed across all records
    """
    changes = []
    columns_changed_overall = set()
    
    # Create unique key (CUST_ID + MOBILE_NO)
    yesterday_df['_key'] = yesterday_df['CUST_ID'].astype(str) + "_" + yesterday_df['MOBILE_NO'].astype(str)
    today_df['_key'] = today_df['CUST_ID'].astype(str) + "_" + today_df['MOBILE_NO'].astype(str)
    
    # For each record in today's worklist
    for idx, today_row in today_df.iterrows():
        cust_id = today_row['CUST_ID']
        today_mobile = today_row['MOBILE_NO']
        
        # Check if this CUST_ID exists in yesterday's worklist
        yesterday_match = yesterday_df[yesterday_df['CUST_ID'] == cust_id]
        
        if len(yesterday_match) == 0:
            # New customer in today's list
            changes.append({
                'CUST_ID': cust_id,
                'CUST_NAME': today_row['CUST_NAME'],
                'CHANGE_TYPE': 'NEW',
                'YESTERDAY_MOBILE_NO': 'N/A',
                'TODAY_MOBILE_NO': today_mobile,
                'YESTERDAY_OFFICE_PH': 'N/A',
                'TODAY_OFFICE_PH': today_row['OFFICE_PH'],
                'YESTERDAY_HOME_PH': 'N/A',
                'TODAY_HOME_PH': today_row['HOME_PH'],
                'YESTERDAY_CUST_NAME': 'N/A',
                'TODAY_CUST_NAME': today_row['CUST_NAME'],
                'COLUMNS_CHANGED_LIST': ['CUST_NAME', 'OFFICE_PH', 'HOME_PH', 'MOBILE_NO']
            })
        else:
            # Customer exists, check for differences
            yesterday_row = yesterday_match.iloc[0]
            
            columns_changed = []
            change_detail = {}
            
            if yesterday_row['CUST_NAME'] != today_row['CUST_NAME']:
                columns_changed.append('CUST_NAME')
                columns_changed_overall.add('CUST_NAME')
                change_detail['YESTERDAY_CUST_NAME'] = yesterday_row['CUST_NAME']
                change_detail['TODAY_CUST_NAME'] = today_row['CUST_NAME']
            else:
                change_detail['YESTERDAY_CUST_NAME'] = yesterday_row['CUST_NAME']
                change_detail['TODAY_CUST_NAME'] = today_row['CUST_NAME']
            
            if str(yesterday_row['OFFICE_PH']) != str(today_row['OFFICE_PH']):
                columns_changed.append('OFFICE_PH')
                columns_changed_overall.add('OFFICE_PH')
                change_detail['YESTERDAY_OFFICE_PH'] = yesterday_row['OFFICE_PH']
                change_detail['TODAY_OFFICE_PH'] = today_row['OFFICE_PH']
            else:
                change_detail['YESTERDAY_OFFICE_PH'] = yesterday_row['OFFICE_PH']
                change_detail['TODAY_OFFICE_PH'] = today_row['OFFICE_PH']
            
            if str(yesterday_row['HOME_PH']) != str(today_row['HOME_PH']):
                columns_changed.append('HOME_PH')
                columns_changed_overall.add('HOME_PH')
                change_detail['YESTERDAY_HOME_PH'] = yesterday_row['HOME_PH']
                change_detail['TODAY_HOME_PH'] = today_row['HOME_PH']
            else:
                change_detail['YESTERDAY_HOME_PH'] = yesterday_row['HOME_PH']
                change_detail['TODAY_HOME_PH'] = today_row['HOME_PH']
            
            if str(yesterday_row['MOBILE_NO']) != str(today_row['MOBILE_NO']):
                columns_changed.append('MOBILE_NO')
                columns_changed_overall.add('MOBILE_NO')
                change_detail['YESTERDAY_MOBILE_NO'] = yesterday_row['MOBILE_NO']
                change_detail['TODAY_MOBILE_NO'] = today_row['MOBILE_NO']
            else:
                change_detail['YESTERDAY_MOBILE_NO'] = yesterday_row['MOBILE_NO']
                change_detail['TODAY_MOBILE_NO'] = today_row['MOBILE_NO']
            
            # Only add to changes if there were actual differences
            if columns_changed:
                changes.append({
                    'CUST_ID': cust_id,
                    'CUST_NAME': today_row['CUST_NAME'],
                    'CHANGE_TYPE': 'UPDATED',
                    'COLUMNS_CHANGED': ', '.join(columns_changed),
                    'COLUMNS_CHANGED_LIST': columns_changed,
                    **change_detail
                })
    
    # Check for deleted customers (in yesterday but not in today)
    for idx, yesterday_row in yesterday_df.iterrows():
        cust_id = yesterday_row['CUST_ID']
        today_match = today_df[today_df['CUST_ID'] == cust_id]
        
        if len(today_match) == 0:
            changes.append({
                'CUST_ID': cust_id,
                'CUST_NAME': yesterday_row['CUST_NAME'],
                'CHANGE_TYPE': 'DELETED',
                'YESTERDAY_MOBILE_NO': yesterday_row['MOBILE_NO'],
                'TODAY_MOBILE_NO': 'N/A',
                'YESTERDAY_OFFICE_PH': yesterday_row['OFFICE_PH'],
                'TODAY_OFFICE_PH': 'N/A',
                'YESTERDAY_HOME_PH': yesterday_row['HOME_PH'],
                'TODAY_HOME_PH': 'N/A',
                'YESTERDAY_CUST_NAME': yesterday_row['CUST_NAME'],
                'TODAY_CUST_NAME': 'N/A',
                'COLUMNS_CHANGED_LIST': ['CUST_NAME', 'OFFICE_PH', 'HOME_PH', 'MOBILE_NO']
            })
    
    changes_df = pd.DataFrame(changes)
    
    # Summary statistics
    if len(changes_df) > 0:
        new_count = len(changes_df[changes_df['CHANGE_TYPE'] == 'NEW'])
        updated_count = len(changes_df[changes_df['CHANGE_TYPE'] == 'UPDATED'])
        deleted_count = len(changes_df[changes_df['CHANGE_TYPE'] == 'DELETED'])
    else:
        new_count = 0
        updated_count = 0
        deleted_count = 0
    
    summary_stats = {
        'total_records_yesterday': len(yesterday_df),
        'total_records_today': len(today_df),
        'new_customers': new_count,
        'updated_customers': updated_count,
        'deleted_customers': deleted_count,
        'total_changes': len(changes_df),
    }
    
    return changes_df, summary_stats, columns_changed_overall


def generate_updated_file(today_df, changes_df, columns_changed_overall):
    """
    Generate ONLY updated records from today's worklist with only changed columns.
    Returns records that have CHANGE_TYPE of 'UPDATED_[COLUMNS]' showing which columns changed.
    """
    # Get only the updated records (exclude NEW, DELETED, NO_CHANGE)
    if len(changes_df) > 0:
        updated_changes = changes_df[changes_df['CHANGE_TYPE'] == 'UPDATED']
        if len(updated_changes) == 0:
            output_df = pd.DataFrame(columns=['CHANGE_TYPE', 'CUST_ID'])
            return output_df
        
        updated_records = updated_changes['CUST_ID'].tolist()
        output_df = today_df[today_df['CUST_ID'].isin(updated_records)].copy()
        
        # Create detailed CHANGE_TYPE showing which columns changed
        change_type_map = {}
        for idx, row in updated_changes.iterrows():
            cust_id = row['CUST_ID']
            columns_list = row['COLUMNS_CHANGED_LIST']
            # Create: UPDATED_COLUMN1_COLUMN2_COLUMN3
            detailed_type = 'UPDATED_' + '_'.join(columns_list)
            change_type_map[cust_id] = detailed_type
        
        output_df['CHANGE_TYPE'] = output_df['CUST_ID'].map(change_type_map)
    else:
        output_df = pd.DataFrame(columns=['CHANGE_TYPE', 'CUST_ID'])
        return output_df
    
    # Format CUST_ID with leading 0: 0&CUST_ID
    output_df['CUST_ID'] = '0' + output_df['CUST_ID'].astype(str)
    
    # Build required columns: CHANGE_TYPE, CUST_ID, plus only the columns that changed
    required_cols = ['CHANGE_TYPE', 'CUST_ID']
    
    # Add only columns that were changed (sorted alphabetically)
    if columns_changed_overall:
        for col in sorted(columns_changed_overall):
            if col in output_df.columns:
                required_cols.append(col)
    
    output_df = output_df[required_cols]
    
    return output_df


# UI Layout
col1, col2 = st.columns(2)

with col1:
    st.subheader("📋 Yesterday's Worklist")
    yesterday_file = st.file_uploader("Upload yesterday's worklist", type=['xlsx', 'xls'], key='yesterday')

with col2:
    st.subheader("📋 Today's Worklist")
    today_file = st.file_uploader("Upload today's worklist", type=['xlsx', 'xls'], key='today')

if yesterday_file and today_file:
    # Load files
    yesterday_df = load_and_validate_file(yesterday_file)
    today_df = load_and_validate_file(today_file)
    
    if yesterday_df is not None and today_df is not None:
        # Compare worklists
        changes_df, summary_stats, columns_changed_overall = compare_worklists(yesterday_df, today_df)
        
        # Display summary statistics
        st.header("📊 Comparison Summary")
        cols = st.columns(5)
        cols[0].metric("Yesterday's Records", summary_stats['total_records_yesterday'])
        cols[1].metric("Today's Records", summary_stats['total_records_today'])
        cols[2].metric("New Customers", summary_stats['new_customers'], delta_color="normal")
        cols[3].metric("Updated Customers", summary_stats['updated_customers'], delta_color="off")
        cols[4].metric("Deleted Customers", summary_stats['deleted_customers'], delta_color="inverse")
        
        # Display changes
        if len(changes_df) > 0:
            st.header("🔄 Detailed Changes")
            
            # Filter by change type
            change_type_filter = st.multiselect(
                "Filter by change type:",
                options=['NEW', 'UPDATED', 'DELETED'],
                default=['NEW', 'UPDATED', 'DELETED']
            )
            
            filtered_changes = changes_df[changes_df['CHANGE_TYPE'].isin(change_type_filter)]
            
            if len(filtered_changes) > 0:
                st.dataframe(
                    filtered_changes,
                    use_container_width=True,
                    height=400,
                    hide_index=True
                )
                
                # Download option for detailed changes
                csv_changes = filtered_changes.to_csv(index=False)
                st.download_button(
                    label="📥 Download Changes Report (CSV)",
                    data=csv_changes,
                    file_name=f"worklist_changes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    mime="text/csv"
                )
            else:
                st.info("No changes match the selected filters.")
        else:
            st.success("✅ No changes detected between the two worklists!")
        
        # Generate updated worklist file - ONLY updated records
        st.header("📝 Updated Records Only")
        updated_df = generate_updated_file(today_df, changes_df, columns_changed_overall)
        
        if len(updated_df) > 0:
            st.subheader(f"Total Updated Records: {len(updated_df)}")
            
            # Show which columns were updated
            if columns_changed_overall:
                cols_str = ", ".join(sorted(columns_changed_overall))
                st.info(f"📊 Columns with changes: **{cols_str}**")
            
            st.dataframe(
                updated_df,
                use_container_width=True,
                height=500,
                hide_index=True
            )
            
            # Export to Excel with formatting
            from io import BytesIO
            from openpyxl.styles import PatternFill
            
            excel_buffer = BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                updated_df.to_excel(writer, sheet_name='Updated Records', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Updated Records']
                
                # Apply yellow highlighting for updated records
                fill_yellow = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                
                for row in worksheet.iter_rows(min_row=2, max_row=len(updated_df) + 1):
                    for cell in row:
                        cell.fill = fill_yellow
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            excel_buffer.seek(0)
            
            st.download_button(
                label="📥 Download Updated Records (Excel)",
                data=excel_buffer,
                file_name=f"updated_records_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            # CSV Export
            csv_export = updated_df.to_csv(index=False)
            st.download_button(
                label="📥 Download Updated Records (CSV)",
                data=csv_export,
                file_name=f"updated_records_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                mime="text/csv"
            )
        else:
            st.info("ℹ️ No updated records found.")

st.info("""
**How it works:**
1. Upload yesterday's worklist
2. Upload today's worklist
3. The system compares them using CUST_ID as the primary identifier
4. Changes in CUST_NAME, OFFICE_PH, HOME_PH, and MOBILE_NO are detected
5. Download the report and updated worklist files
""")
